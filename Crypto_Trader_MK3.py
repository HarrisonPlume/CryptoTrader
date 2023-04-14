import sys 
sys.path.append('/usr/lib/python3/dist-packages')
import requests
import xlsxwriter
import json
import ast
import time
import pandas as pd
import openpyxl
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart

def CreateHeader():
    # Create an new Excel file and add a worksheet.
    workbook = xlsxwriter.Workbook('Finance.xlsx')
    worksheet = workbook.add_worksheet()

    # Increase the cell size of the merged cells to highlight the formatting.
    worksheet.set_column('A:G', 15)


    # Create a format to use in the merged range.
    merge_format = workbook.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '#acacac'})


    # Merge 3 cells.
    worksheet.merge_range('A1:C1', 'Crypto Trader MK2', merge_format)
    worksheet.write("A2","Inital Funds")
    worksheet.write("B2",100)
    worksheet.write("A3","Asset id")
    worksheet.write("B3","Asset Code")
    worksheet.write("C3","Asset Buy Price")
    worksheet.write("D3","Asset Sell Price")
    worksheet.write("E3","Asset purchae amount")
    worksheet.write("F3","Funds at Purchase")
    worksheet.write("G3","Funds at Sale")
    worksheet.write("H3","time")
    worksheet.write("I3","Profit")

    workbook.close()


def get_market_assets():
    response = requests.get('https://api.swyftx.com.au/markets/assets/')

    # raw_assets = response.text.split("},")
    raw_assets = json.loads(response.text)
    refined_assets = []

    for asset in raw_assets:
        newasset = {}
        if asset["isRestricted"] == False or asset["tradable"] == 1 or asset["buyDisabled"] == 0:
            newasset["id"] = asset["id"]
            newasset["code"] = asset["code"]
            if newasset["code"] == "AUD":
                continue
            elif newasset["code"] == "SUB":
                continue
            elif newasset["code"] == "SALT":
                continue
            elif newasset["code"] == "MCO":
                continue
            elif newasset["code"] == "GBP":
                continue
            elif newasset["code"] == "JPY":
                continue
            elif newasset["code"] == "EUR":
                continue
            elif newasset["code"] == "USD":
                continue
            elif newasset["code"] == "BTT":
                continue
            elif newasset["code"] == "NANO":
                continue
            elif newasset["code"] == "NPXS":
                continue
            elif newasset["code"] == "PPT":
                continue
            elif newasset["code"] == "USDT":
                continue
            else:
                refined_assets.append(newasset)
        else:
            continue

    return refined_assets

def get_asset_price(code):
    """
    function gets the current market price of the asset

    """
    response = requests.get('https://api.swyftx.com.au/markets/info/basic/{}/'.format(code))
    response = response.text[1:]
    response = response[:-1]
    try:
        result = json.loads(response)
    except:
        return float(0.0), float(0.0)
    try:
        buy_price = result["buy"]
    except:
        return float(0.0), float(0.0)

    return float(result["buy"]), float(result["sell"])

def get_highest_apreciating_asset_first():
    timestart = time.time()
    response = requests.get('https://api.swyftx.com.au/live-rates/1/')

    #Get data frame 1
    responser = response.text
    prices_by_id = json.loads(responser)

    data_frame_1 = pd.DataFrame.from_dict(prices_by_id).transpose().drop(['1'], axis=0)
    data_frame_1["midPrice"] = data_frame_1['midPrice'].astype(float)
    data_frame_1 = data_frame_1[data_frame_1.midPrice != 0]
    # print(data_frame_1['midPrice'])

    time.sleep(0.2)

    response = requests.get('https://api.swyftx.com.au/live-rates/1/')
    responser = response.text
    prices_by_id = json.loads(responser)
    data_frame_2 = pd.DataFrame.from_dict(prices_by_id).transpose().drop(['1'], axis=0)
    data_frame_2["midPrice"] = data_frame_2['midPrice'].astype(float)
    data_frame_2 = data_frame_2[data_frame_2.midPrice != 0]


    price_diffs = pd.DataFrame(columns=["id","MPC"])
    counter = 0
    # assert len(data_frame_1) == len(data_frame_2), "Price Dataframes differ in length stopping program..."
    for midprice in data_frame_1['midPrice']:
        index = str(data_frame_1[data_frame_1["midPrice"]==midprice].index.values[0])
        try:
            price_difference_percent  = float(((data_frame_2['midPrice'].loc[index]-midprice)/((data_frame_2['midPrice'].loc[index]+midprice)/2))*100)
        except:
            price_difference_percent = 0.0
        price_diffs.loc[counter] = [index]+[price_difference_percent]
        counter += 1
    price_diffs = price_diffs.sort_values(by = "MPC", ascending=False)
    print('\n')
    print(price_diffs.head())


    topAsset = price_diffs.iloc[0]
    return topAsset['id'], topAsset["MPC"]

def get_highest_apreciating_asset():
    timestart = time.time()
    response = requests.get('https://api.swyftx.com.au/live-rates/1/')

    #Get data frame 1
    responser = response.text
    prices_by_id = json.loads(responser)
    data_frame_1 = pd.DataFrame.from_dict(prices_by_id).transpose().drop(['1'], axis=0)
    data_frame_1["midPrice"] = data_frame_1['midPrice'].astype(float)
    data_frame_1 = data_frame_1[data_frame_1.midPrice != 0]
    # print(data_frame_1['midPrice'])

    # time.sleep(0.1)

    response = requests.get('https://api.swyftx.com.au/live-rates/1/')
    responser = response.text
    prices_by_id = json.loads(responser)
    data_frame_2 = pd.DataFrame.from_dict(prices_by_id).transpose().drop(['1'], axis=0)
    data_frame_2["midPrice"] = data_frame_2['midPrice'].astype(float)
    data_frame_2 = data_frame_2[data_frame_2.midPrice != 0]


    price_diffs = pd.DataFrame(columns=["id","MPC"])
    counter = 0
    # assert len(data_frame_1) == len(data_frame_2), "Price Dataframes differ in length stopping program..."
    for midprice in data_frame_1['midPrice']:
        index = str(data_frame_1[data_frame_1["midPrice"]==midprice].index.values[0])
        try:
            price_difference_percent  = float(((data_frame_2['midPrice'].loc[index]-midprice)/((data_frame_2['midPrice'].loc[index]+midprice)/2))*100)
        except:
            price_difference_percent  = 0
        price_diffs.loc[counter] = [index]+[price_difference_percent]
        counter += 1
    price_diffs = price_diffs.sort_values(by = "MPC", ascending=False)
    

    topAsset = price_diffs.iloc[0]
    return topAsset['id'], topAsset["MPC"]

def RunAlg():
    Funds = 100
    print("Starting... ${} AUD".format(Funds))
    timestart = time.time()
    #Get assets
    rawAssetList = pd.DataFrame.from_dict(get_market_assets())
    print("Getting market assets took:",float(time.time()-timestart),"s")
    timestart = time.time()
    #get higest apreciating asset at the moment  n
    newAssetPerformace = 0
    while newAssetPerformace < 0.3:
        new_assetid, newAssetPerformace = get_highest_apreciating_asset_first()
    print("Getting Higest performing asset took:",float(time.time()-timestart),"s")
    timestart = time.time() 
    newassetcode = rawAssetList.query("id=={}".format(str(new_assetid)))
    try:
        newassetcode = newassetcode.iloc[0, 1]
    except:
        print("cooked")
        new_assetid, newAssetPerformace = get_highest_apreciating_asset_first()
        newassetcode = rawAssetList.query("id=={}".format(str(new_assetid)))
        newassetcode = newassetcode.iloc[0, 1]
    print("new_asset:",newassetcode)
    print("Selecting HAA took:",float(time.time()-timestart),"s")
    timestart = time.time()
    # #Get asset price
    buy, sell = get_asset_price(newassetcode)
    currentAssetBuy = buy
    #Create Asset object to be put into excel spreadsheet
    currentAsset = pd.DataFrame(columns=["id","code","Buy","Sell","Amount","FP"])
    #Buy seleted asset with a 0.6% brokerage fee
    Asset_Amount = Funds/(buy*1.006)
    Funds_at_purchase = Asset_Amount*buy
    currentAsset.loc[0] = [str(new_assetid)]+[str(newassetcode)]+[float(buy)]+[float(sell)]+[float(Asset_Amount)]+[float(Funds_at_purchase)]
    print(currentAsset)
    currentAssetCode = str(newassetcode)
    print("Purchase of current asset took:",float(time.time()-timestart),"s")
    with pd.ExcelWriter('Finance.xlsx',engine="openpyxl", mode='a',if_sheet_exists="overlay") as writer:
        currentAsset.to_excel(writer, sheet_name='Sheet1', index=False, header = False, startrow=writer.sheets['Sheet1'].max_row )

    #Profit tracker
    oldProfit_percent = 0
    #Run code to keep checking highest performing asset and swap to it when required.
    counter = 0
    while True:
        #get higest apreciating asset at the moment
        # new_assetid, newAssetPerformace = get_highest_apreciating_asset()
        if new_assetid != str(currentAsset["id"]):
            #sell current asset
            buy, sell = get_asset_price(currentAssetCode)
            if_sold_Funds = Asset_Amount*(sell-sell*0.006)
            Profit = if_sold_Funds - Funds
            Profit_percent  = float(((if_sold_Funds-Funds)/((if_sold_Funds+Funds)/2))*100)
            if Profit < 0:
                print("\n")
                print("Asset Name:",newassetcode)
                print("Buy Price:",currentAssetBuy)
                print("Sell Price:",sell)
                print("Profit:",Profit)
                if Profit_percent < -4.5:
                    #Sell condition
                    print("SOLD!, Negative")
                    print(Profit_percent)
                    Funds = if_sold_Funds
                    currentAsset["FS"] = Funds
                    with pd.ExcelWriter('Finance.xlsx',engine="openpyxl", mode='a',if_sheet_exists="overlay") as writer:
                        currentAsset.to_excel(writer, sheet_name='Sheet1', index=False, header = False, startrow=writer.sheets['Sheet1'].max_row )
                    break
            else:
                if Profit_percent > oldProfit_percent:   
                    #Check to see if the item is going negative
                    print("\n")
                    print("Asset Name:",newassetcode)
                    print("Buy Price:",currentAssetBuy)
                    print("Sell Price:",sell)
                    print("Profit:",Profit)
                    oldProfit_percent = Profit_percent
                else:
                    
                    #Sell condition
                    print("SOLD!")
                    print("Positive Profit!")
                    Funds = if_sold_Funds
                    currentAsset["FS"] = Funds
                    # Set up the SMTP server details for Outlook
                    smtp_server = "smtp.office365.com"
                    smtp_port = 587
                    smtp_username = "plume521@outlook.com"
                    smtp_password = "Tq49XUQ7"

                    # Set up the email details
                    sender = "plume521@outlook.com"
                    recipient = "plume521@outlook.com"
                    subject = "Crypto trader recorded a positive Profit"
                    message = "Hi Haz, a possitive proffit was recorded on your crypto trader bot. {} sold for: {}".format(newassetcode,if_sold_Funds)

                    # Create a MIME message
                    msg = MIMEMultipart()
                    msg['From'] = sender
                    msg['To'] = recipient
                    msg['Subject'] = subject
                    msg.attach(MIMEText(message))

                    # Connect to the SMTP server and send the email
                    with smtplib.SMTP(smtp_server, smtp_port) as server:
                        server.starttls()
                        server.login(smtp_username, smtp_password)
                        server.sendmail(sender, recipient, msg.as_string())
                    with pd.ExcelWriter('Finance.xlsx',engine="openpyxl", mode='a',if_sheet_exists="overlay") as writer:
                        currentAsset.to_excel(writer, sheet_name='Sheet1', index=False, header = False, startrow=writer.sheets['Sheet1'].max_row )
                    break


    print(currentAsset)



CreateHeader()
RunAlg()
import sys 
sys.path.append('/usr/lib/python3/dist-packages')
import requests
import xlsxwriter
import json
import ast
import time
import pandas as pd
import openpyxl

def CreateHeader():
    # Create an new Excel file and add a worksheet.
    workbook = xlsxwriter.Workbook('Finance.xlsx')
    worksheet = workbook.add_worksheet()

    # Increase the cell size of the merged cells to highlight the formatting.
    worksheet.set_column('A:G', 15)


    # Create a format to use in the merged range.
    merge_format = workbook.add_format({
        'bold': 1,
        'border': 1,
        'align': 'center',
        'valign': 'vcenter',
        'fg_color': '#acacac'})


    # Merge 3 cells.
    worksheet.merge_range('A1:C1', 'Crypto Trader MK2', merge_format)
    worksheet.write("A3","Asset Code")
    worksheet.write("A2","Inital Funds")
    worksheet.write("B2",100)
    worksheet.write("B3","Asset Buy Price")
    worksheet.write("C3","Asset Buy Diff (%)")
    worksheet.write("D3","Asset Sell Price")
    worksheet.write("E3","Asset Sell Diff (%)")
    worksheet.write("F3","Asset purchae amount")
    worksheet.write("G3","Funds at Purchase")
    worksheet.write("H3","Funds at Sale")
    worksheet.write("I3","time")
    worksheet.write("J3","Profit")

    workbook.close()

def Get_Authorisation():
    values = '''{
        "apikey": "f_iQ_FrQWbKYuSMsFJbwpYvuz45mfgjvA-Ahip-Ifog0P"
        }'''
    
    headers = {
        'Content-Type': 'application/json'
    }

    response = requests.post('https://api.swyftx.com.au/auth/refresh/',values,headers)
    # print(response.status_code)
    # print(response.headers)
    print(response.json())

token = "eyJhbGciOiJSUzI1NiIsInR5cCI6IkpXVCIsImtpZCI6IlJrVTRRelF6TlRaQk5rTkNORGsyTnpnME9EYzNOVEZGTWpaRE9USTRNalV6UXpVNE1UUkROUSJ9.eyJodHRwczovL3N3eWZ0eC5jb20uYXUvLWp0aSI6IjZmY2I0ODY4LWQ2NzAtNDE1Yi05MjcyLTU3ZGVmZTZiOTgwNSIsImh0dHBzOi8vc3d5ZnR4LmNvbS5hdS8tbWZhX2VuYWJsZWQiOmZhbHNlLCJodHRwczovL3N3eWZ0eC5jb20uYXUvLXVzZXJVdWlkIjoidXNyX01QRTFGVnRrUDc4MW5jYjQzajJyTkwiLCJodHRwczovL3N3eWZ0eC5jb20uYXUvLWNvdW50cnlfbmFtZSI6IkF1c3RyYWxpYSIsImh0dHBzOi8vc3d5ZnR4LmNvbS5hdS8tY2l0eV9uYW1lIjoiU3lkbmV5IiwiaXNzIjoiaHR0cHM6Ly9zd3lmdHguYXUuYXV0aDAuY29tLyIsInN1YiI6ImF1dGgwfDY0MWZhNmQ2NmFjNTg0YmUzMWJmYzQ0NCIsImF1ZCI6Imh0dHBzOi8vYXBpLnN3eWZ0eC5jb20uYXUvIiwiaWF0IjoxNjc5ODA1MTIxLCJleHAiOjE2ODA0MDk5MjEsImF6cCI6IkVRdzNmYUF4T1RoUllUWnl5MXVsWkRpOERIUkFZZEVPIiwic2NvcGUiOiJhcHAgYXBwLmFjY291bnQgYXBwLmFjY291bnQuYWZmaWxpYXRpb24gYXBwLmFjY291bnQubW9kaWZ5IGFwcC5hY2NvdW50LnRheC1yZXBvcnQgYXBwLmFjY291bnQudmVyaWZpY2F0aW9uIGFwcC5hY2NvdW50LmJhbGFuY2UgYXBwLmFjY291bnQuc3RhdHMgYXBwLmFjY291bnQucmVhZCBhcHAucmVjdXJyaW5nLW9yZGVycyBhcHAucmVjdXJyaW5nLW9yZGVycy5yZWFkIGFwcC5yZWN1cnJpbmctb3JkZXJzLmNyZWF0ZSBhcHAucmVjdXJyaW5nLW9yZGVycy5kZWxldGUgYXBwLmFkZHJlc3MgYXBwLmFkZHJlc3MuYWRkIGFwcC5hZGRyZXNzLnJlbW92ZSBhcHAuYWRkcmVzcy5jaGVjay1kZXBvc2l0IGFwcC5hZGRyZXNzLnJlYWQgYXBwLmZ1bmRzIGFwcC5mdW5kcy53aXRoZHJhdyBhcHAuZnVuZHMud2l0aGRyYXdhbC1saW1pdCBhcHAuZnVuZHMucmVhZCBhcHAub3JkZXJzIGFwcC5vcmRlcnMuY3JlYXRlIGFwcC5vcmRlcnMuZGVsZXRlIGFwcC5vcmRlcnMucmVhZCBhcHAub3JkZXJzLmR1c3QgYXBwLmFwaSBhcHAuYXBpLnJldm9rZSBhcHAuYXBpLnJlYWQgb2ZmbGluZV9hY2Nlc3MiLCJndHkiOiJwYXNzd29yZCJ9.c_QCocppxx-VccdJAe2RvYvqfm6doqRuRmeLj-olHHwsFPFgJfmY84d6RkE54sErplkolMzOm9GMtQ37Rt9glNsXdHSUxMOO2NVB2r61aL0sM3Sl7qJzudE6oNk6OFOvLnsXWxP6i8OpqJLhYRbhVdL3M73OksbqQhL-eH5lTz_bmhSmz1sID1Xc-u91_SIbWj6jjl5Ox8PNLs6N-saUpphaa6-LIh15zPc2-zCGRIm5JsOAd0BSGHe_jD3gAc3UZtmpCnnEYFwvnFMyvUESWzaeka70TO8Opi9-lr_7_xgQuvPIntIpl0xNjvezkGg9wvaRciXDnmAJp9XdvlA85Q"

def Get_account_info():
    
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer {}'.format(token)
    }

    response = requests.get('https://api.swyftx.com.au/user/', headers=headers)

    print(response.status_code)
    print(response.json())

def Get_account_balance():
    headers = {
    'Content-Type': 'application/json',
    'Authorization': 'Bearer {}'.format(token)
    }

    response = requests.get('https://api.swyftx.com.au/user/balance/', headers=headers)

    print(response.status_code)
    print(response.json())


def get_market_assets():
    response = requests.get('https://api.swyftx.com.au/markets/assets/')

    # raw_assets = response.text.split("},")
    raw_assets = json.loads(response.text)
    refined_assets = []

    for asset in raw_assets:
        newasset = {}
        if asset["isRestricted"] == False or asset["tradable"] == 1 or asset["buyDisabled"] == 0:
            newasset["code"] = asset["code"]

            if newasset["code"] == "AUD":
                continue
            elif newasset["code"] == "SUB":
                continue
            elif newasset["code"] == "SALT":
                continue
            elif newasset["code"] == "MCO":
                continue
            elif newasset["code"] == "GBP":
                continue
            elif newasset["code"] == "JPY":
                continue
            elif newasset["code"] == "EUR":
                continue
            elif newasset["code"] == "USD":
                continue
            elif newasset["code"] == "BTT":
                continue
            elif newasset["code"] == "NANO":
                continue
            elif newasset["code"] == "NPXS":
                continue
            elif newasset["code"] == "PPT":
                continue
            elif newasset["code"] == "USDT":
                continue
            else:
                refined_assets.append(newasset)
        else:
            continue

    return refined_assets

def get_new_asset(dataframe):

    return Asset_code

    
def get_asset_price(code):
    """
    function gets the current market price of the asset

    """
    response = requests.get('https://api.swyftx.com.au/markets/info/basic/{}/'.format(code))
    response = response.text[1:]
    response = response[:-1]
    try:
        result = json.loads(response)
    except:
        return float(0.0), float(0.0)
    try:
        buy_price = result["buy"]
    except:
        return float(0.0), float(0.0)

    return float(result["buy"]), float(result["sell"])


def RunAlg():
    Funds = 100
    print("Starting... ${} AUD".format(Funds))
    timestart = time.time()
    rawAssetList = get_market_assets()
    print("Getting market assets took:",float(time.time()-timestart),"s")
    timestart = time.time()
    refinedAssetList = []
    #Get inital prices
    for Asset in rawAssetList:
        buy, sell = get_asset_price(Asset['code'])
        #Remove assets that have a zero buy price
        if buy != 0:
            Asset["buy"] = float(buy)
            Asset["sell"] = float(sell)
            refinedAssetList.append(Asset)
    print("Asset Prices took:",float(time.time()-timestart),"s")
    # print(refinedAssetList)

    #Create  a loop that checks the new asset price with the old asset price
    newAssets = pd.DataFrame(columns=["code","buy","BC","sell","SC"])
    timestart = time.time()
    counter = 0
    for Asset in refinedAssetList:
        newAsset = pd.DataFrame
        buy, sell = get_asset_price(Asset['code'])
        buy_price_diff = ((buy - Asset["buy"])/((buy + Asset["buy"])/2))*100
        sell_price_diff = ((sell - Asset["sell"])/((sell+Asset['sell'])/2))*100
        newAssets.loc[counter] = [Asset["code"]]+[float(buy)]+[float(buy_price_diff)]+[float(sell)]+[float(sell_price_diff)]
        counter += 1
    print("Refined dataframe took:",float(time.time()-timestart),"s")

    #Asset selection
    newAssets = newAssets.sort_values(by="BC", ascending=False)
    # print(newAssets)

    #select asset with highest sell price
    Current_Asset = newAssets.iloc[0]
    Current_Asset = Current_Asset.to_frame().transpose()
    
    #Buy seleted asset with a 0.6% brokerage fee
    Asset_Amount = Funds/(Current_Asset["buy"]*1.006)
    Current_Asset["amount"] = Asset_Amount
    Funds_at_purchase = Asset_Amount*Current_Asset["buy"]
    Current_Asset["Purchase_funds"] = Funds_at_purchase
    Current_Asset_code = str(Current_Asset["code"])
    print("Current_asset",Current_Asset_code)

    ####
    #Perform buy operation
    ####

    #Write First asset purchase
    with pd.ExcelWriter('Finance.xlsx',engine="openpyxl", mode='a',if_sheet_exists="overlay") as writer:
        Current_Asset.to_excel(writer, sheet_name='Sheet1', index=False, header = False, startrow=writer.sheets['Sheet1'].max_row )

    #Check asset price at timestep
    counter = 0
    while counter < 1:
        asset_counter = 0
        timestart = time.time()
        for asset in newAssets:
            buy, sell = get_asset_price(Asset['code'])
            buy_price_diff = ((buy - Asset["buy"])/((buy + Asset["buy"])/2))*100
            sell_price_diff = ((sell - Asset["sell"])/((sell+Asset['sell'])/2))*100
            newAssets.loc[asset_counter] = [Asset["code"]]+[float(buy)]+[float(buy_price_diff)]+[float(sell)]+[float(sell_price_diff)]
            asset_counter += 1
        print("Refined dataframe took:",float(time.time()-timestart),"s")
        newAssets = newAssets.sort_values(by="BC", ascending=False)
        new_Asset = newAssets.iloc[0]
        new_Asset = new_Asset.to_frame().transpose()
        print(new_Asset)
        New_Asset_code = str(new_Asset["code"])
        print("New_Asset",New_Asset_code)
        if New_Asset_code != Current_Asset_code:
            #Sell current asset and buy another
            buy, sell = get_asset_price(Current_Asset_code)
            Funds = Current_Asset["amount"]*(sell-sell*0.006)
            #Update spreasheet with sale amount
            Current_Asset["Sale_funds"] = Funds
            with pd.ExcelWriter('Finance.xlsx',engine="openpyxl", mode='w',if_sheet_exists="overlay") as writer:
                Current_Asset.to_excel(writer, sheet_name='Sheet1', index=False, header = False, startrow=writer.sheets['Sheet1'].max_row )

            # #Buy new asset
            # Asset_Amount = Funds/(Current_Asset["buy"]*1.006)
            # print(Asset_Amount)
            # Current_Asset["amount"] = Asset_Amount
            # Funds_at_purchase = Asset_Amount*(buy-(buy*0.006))
            # Current_Asset["Purchase_funds"] = Funds_at_purchase
            
            # with pd.ExcelWriter('Finance.xlsx',engine="openpyxl", mode='a',if_sheet_exists="overlay") as writer:
            #     Current_Asset.to_excel(writer, sheet_name='Sheet1', index=False, header = False, startrow=writer.sheets['Sheet1'].max_row )

        counter += 1
        




    
CreateHeader()
RunAlg()